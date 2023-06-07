import sys
import openpyxl
import os
import time

"""
@Auther : ga0weI
@Function : get specific column data form xlsx to txt
"""

def findcolnumbyname(name,ws):
    for clnum in range(1, ws.max_column + 1):
        if name in ws.cell(row=1, column=clnum).value:
            return clnum

if __name__ == '__main__':
    """
    两个参数，filename + cloumnname
    """
    filename = sys.argv[1]
    filepath = os.path.join(os.path.abspath("."), filename)
    filecloum = sys.argv[2]
    workbook = openpyxl.load_workbook(filepath)
    ws = workbook.worksheets[0]
    data_count = findcolnumbyname(filecloum, ws)

    maxcolum = ws.max_column + 1
    with open('data.txt','w') as f:
        for i in range(2, ws.max_row + 1):
            datas = ws.cell(row=i, column=data_count).value
            if datas is not None:
                datas =datas.replace('\n','')
                f.write(datas+"\n")

    print("从excel里面提取datas成功，并记录在data.txt文件中")



